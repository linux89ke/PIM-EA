import logging
import os
import re

import pandas as pd
import streamlit as st

logger = logging.getLogger(__name__)


def _clean_text(val) -> str:
    if val is None:
        return ""
    s = str(val)
    return "" if s.strip().lower() in ("nan", "none") else s.strip()


# _kenya_book_mask lived here. It decided whether a product was in a book
# category with `"book" in CATEGORY`, which is wrong in both directions:
# the audit record carries the leaf, so a book correctly filed under
# "Books, Movies and Music / Business & Finance / Business & Economics"
# arrived as "Business & Economics" and failed; and "Nursery Decor /
# Bookends" contains "book" while being no kind of book category.
#
# Books are verified like every other category now. The exemption below is
# the surviving definition, and it splits the path properly.


# ── Kenya: books are exempt from the Wrong Category matcher ─────────────────
#
# The category matcher is a TF-IDF/ML predictor. Book titles are arbitrary
# prose, so it routinely predicts some unrelated category for a book that is
# already filed correctly and the product gets rejected for Wrong Category.
#
# Real taxonomy (from category_map.xlsx):
#     Books, Movies and Music
#     Books, Movies and Music / Art & Humanities / ...
#     Books, Movies and Music / Bestselling Books
#     Books, Movies and Music / DVDs
#     Books, Movies and Music / DVDs / Drama          <- these still get checked
#
# So: anything under the Books branch is exempt, except the DVDs sub-tree.
#
# Matching is prefix-based on path segments, not a substring search, because
# the taxonomy contains traps in both directions:
#     "Baby Products / Nursery / Nursery Decor / Bookends"      contains "book"
#     "Automobile / ... / In-Dash DVD & Video Receivers"        contains "dvd"
# Neither belongs to the Books branch.
_KE_BOOKS_ROOT = "books, movies and music"
_KE_DVD_SEGMENT = "dvds"


def _split_path(path: str) -> list:
    raw = str(path or "").replace(">", "/")
    return [seg.strip().lower() for seg in raw.split("/") if seg.strip()]


def is_kenya_books_exempt(path: str) -> bool:
    """True when a Kenya category path must not be flagged as Wrong Category."""
    segs = _split_path(path)
    if not segs or segs[0] != _KE_BOOKS_ROOT:
        return False
    # Under the Books branch. Only the DVDs sub-tree stays in scope.
    return not (len(segs) > 1 and segs[1] == _KE_DVD_SEGMENT)


def drop_kenya_books_false_positives(
    flagged: pd.DataFrame, code_to_path: dict = None
) -> pd.DataFrame:
    """Remove Kenya book rows the category matcher flagged by mistake."""
    if flagged is None or flagged.empty:
        return flagged

    paths = None
    if "CATEGORY" in flagged.columns:
        paths = flagged["CATEGORY"].astype(str)
    if code_to_path and "CATEGORY_CODE" in flagged.columns:
        _mapped = flagged["CATEGORY_CODE"].apply(
            lambda c: code_to_path.get(str(c).strip(), "") if pd.notna(c) else ""
        )
        # Prefer the mapped path, fall back to whatever CATEGORY held.
        paths = _mapped.where(_mapped.astype(str).str.strip().ne(""), paths) \
            if paths is not None else _mapped
    if paths is None:
        return flagged

    keep = ~paths.apply(is_kenya_books_exempt)
    return flagged[keep].copy()


@st.cache_data(ttl=3600)
def load_health_beauty_codes() -> set:
    try:
        from openpyxl import load_workbook
        cat_wb = load_workbook('category_map.xlsx', read_only=True)
        cat_ws = cat_wb.active
        cat_rows = list(cat_ws.iter_rows(values_only=True))
        cat_wb.close()
        
        # Column 1 = code, Column 2 = path
        hb_codes = set()
        for r in cat_rows[1:]:
            if r[1] and r[2]:
                path = str(r[2]).strip()
                if path.startswith("Health") and "Beauty" in path:
                    hb_codes.add(str(int(r[1])))
        return hb_codes
    except Exception:
        logger.exception("load_health_beauty_codes: failed to load category_map.xlsx — Health & Beauty checks will see an empty code set")
        return set()


# ALLOWLIST: only category path sections that are relevant for KEBS skin-product checking.
# This is MUCH more precise than an exclusion list — only skin care / body beauty / cosmetics
# categories are checked. Hair care, health care, oral care, vision, sexual wellness, 
# supplements, tools, accessories etc. are all naturally excluded.
_KEBS_ALLOWED_PATH_SECTIONS = (
    # Core skin & body care
    "skin care",
    "body care",
    "bathing & skin care",
    "bath & body",
    "body & bath",
    "bath, body & accessories",
    "soaps",
    "cleansers",
    "toners",
    "moisturizers",
    "lotions",
    "face care",
    "body lotion",
    "body scrub",
    "exfoliants",
    "sunscreen",
    "sunblock",
    "sun care",
    "sun protection",
    # Cosmetics / makeup
    "makeup",
    "cosmetics",
    "face makeup",
    "lip makeup",
    "eye makeup",
    "bb & cc cream",
    "foundation",
    "concealer",
    "primer",
    # Fragrances
    "fragrances",
    "perfumes",
    # Men's grooming (skin-related)
    "men's grooming",
    # Dermocosmetics
    "dermocosmetics",
    # Beauty & personal care top-level section
    "beauty & personal care",
    # Baby skin / body care (specifically bathing and skin)
    "bathing & skin care",
    "diapering",
    # Kid's beauty
    "kid's beauty",
    # Deodorants (can contain lightening agents)
    "deodorants",
    "antiperspirants",
    # Shave & hair removal (relevant — depilatory products can have banned ingredients)
    "shave & hair removal",
)


@st.cache_data(ttl=3600)
def load_kebs_hb_codes() -> set:
    """
    Load category codes for KEBS skin-product check using an ALLOWLIST approach.
    Only includes beauty/skin/body care, cosmetics, fragrances, and baby skin care categories.
    All other H&B sections (hair care, oral care, health care, supplements, 
    vision, sexual wellness, tools, etc.) are naturally excluded.
    """
    try:
        from openpyxl import load_workbook
        cat_wb = load_workbook('category_map.xlsx', read_only=True)
        cat_ws = cat_wb.active
        cat_rows = list(cat_ws.iter_rows(values_only=True))
        cat_wb.close()

        kebs_codes = set()
        allowed_lower = tuple(s.lower() for s in _KEBS_ALLOWED_PATH_SECTIONS)

        for r in cat_rows[1:]:
            if not r[1] or not r[2]:
                continue
            path = str(r[2]).strip().lower()

            # Must be in Health & Beauty OR Baby Products (for baby skin care)
            is_hb = path.startswith("health & beauty") or path.startswith("health and beauty")
            is_baby = path.startswith("baby")

            if not (is_hb or is_baby):
                continue

            # Check if any allowed section appears in the path
            if any(section in path for section in allowed_lower):
                try:
                    kebs_codes.add(str(int(r[1])))
                except (ValueError, TypeError):
                    pass

        return kebs_codes
    except Exception:
        logger.exception("load_kebs_hb_codes: failed to load category_map.xlsx — KEBS skin-product check will see an empty code set")
        return set()


@st.cache_data(ttl=3600)
def load_kebs_banned_products():
    """
    Returns two structures:
    1. full_products: list of dicts with {full_name_lower, brand_lower, reason}
       for entries where the Product Name is more than just the brand name.
    2. brand_only: list of dicts with {brand_lower, reason}
       for entries where Product Name == Brand (brand-only entries like 'Elegance').
    """
    xlsx_path = "kebs_banned_products (1).xlsx"
    csv_path = "kebs_banned_products (1).csv"
    try:
        if os.path.exists(xlsx_path):
            df = pd.read_excel(xlsx_path, sheet_name=0, dtype=str)
        elif os.path.exists(csv_path):
            df = pd.read_csv(csv_path, encoding="cp1252", dtype=str)
        else:
            return [], []

        df = df.fillna('')
        full_products = []  # entries with a full product name (longer than just brand)
        brand_only = []     # entries where product name == brand or is very short

        for _, row in df.iterrows():
            reason = str(row.get('Reason for Ban', '')).strip()
            brand = str(row.get('Brand', '')).strip()
            product = str(row.get('Product Name', '')).strip()
            if not brand or brand.lower() in ('nan', 'none', ''):
                continue

            brand_l = brand.lower()
            product_l = product.lower() if product and product.lower() not in ('nan', 'none', '') else ''

            # If the product name is essentially just the brand name or empty,
            # store as brand-only (requires skin-type word in name to match).
            # If it has meaningful words beyond just the brand, store as full product.
            if not product_l or product_l == brand_l or product_l.strip() == brand_l.strip():
                brand_only.append({"brand_lower": brand_l, "reason": reason})
            else:
                full_products.append({
                    "full_name_lower": product_l,
                    "brand_lower": brand_l,
                    "reason": reason,
                })
                # Also add a brand-only entry so brand alone can still match
                brand_only.append({"brand_lower": brand_l, "reason": reason})

        return full_products, brand_only
    except Exception:
        logger.exception(f"load_kebs_banned_products: failed to load {xlsx_path!r}/{csv_path!r} — KEBS banned-product check will see an empty list")
        return [], []


# Skin-lightening product type indicators — if a product name contains any of these
# it is considered a skin-lightening product and will be checked against the brand list.
_SKIN_PRODUCT_TYPES = re.compile(
    r"\b(?:cream|creme|lotion|gel|soap|serum|milk|toner|bleach|balm|oil|lemon|lait|lait corp|"
    r"fade|skin|lightening|whitening|brightening|complexion|beauty|medicated|clarifying|"
    r"moisturiz|moisturis|body|face|treatment)\b",
    re.IGNORECASE,
)


def check_kebs_banned_products(data: pd.DataFrame) -> pd.DataFrame:
    """
    Checks uploaded products against the KEBS banned products list.

    Matching strategy (all three common upload patterns are covered):
      Case A - Generic brand / full product name:
        BRAND='Generic', NAME='Elegance Skin Lightening Cream'
        → BRAND+NAME = 'generic elegance skin lightening cream'
        → Full product name 'elegance skin lightening cream' found as substring → MATCH

      Case B - Brand in BRAND, product type only in NAME:
        BRAND='Elegance', NAME='Skin Lightening Cream'
        → Combined = 'elegance skin lightening cream'
        → Full product name found → MATCH

      Case C - Brand repeated in both:
        BRAND='Elegance', NAME='Elegance Skin Lightening Cream'
        → Combined = 'elegance elegance skin lightening cream'
        → Full product name found → MATCH

      Brand-only entries (e.g., 'Elegance' with Product Name == 'Elegance'):
        → BRAND matches AND NAME contains a skin-product type word → MATCH

    Only applies to Health & Beauty category codes.
    """
    required = {"PRODUCT_SET_SID", "NAME", "BRAND"}
    if data.empty or not required.issubset(data.columns):
        return pd.DataFrame(columns=data.columns)

    full_products, brand_only = load_kebs_banned_products()
    if not full_products and not brand_only:
        return pd.DataFrame(columns=data.columns)

    hb_codes = load_kebs_hb_codes()
    if not hb_codes:
        return pd.DataFrame(columns=data.columns)

    d = data.copy()

    # Filter for H&B categories only
    mask_hb = pd.Series(False, index=d.index)
    if "CATEGORY_CODE" in d.columns:
        mask_hb = d["CATEGORY_CODE"].astype(str).str.strip().isin(hb_codes)
    
    if not mask_hb.all() and "CATEGORY" in d.columns:
        cat_lower = d["CATEGORY"].astype(str).str.lower()
        fallback_mask = pd.Series(False, index=d.index)
        for allowed in _KEBS_ALLOWED_PATH_SECTIONS:
            fallback_mask |= cat_lower.str.contains(allowed.lower(), na=False)
        mask_hb = mask_hb | fallback_mask
        
    target = d[mask_hb].copy()
    if target.empty:
        return pd.DataFrame(columns=data.columns)

    # Build combined text: "brand name" for matching
    target["_brand_l"] = target["BRAND"].fillna("").str.strip().str.lower()
    target["_name_l"]  = target["NAME"].fillna("").str.strip().str.lower()
    # Combined = brand + " " + name (handles Case A, B, C)
    target["_combined"] = (target["_brand_l"] + " " + target["_name_l"]).str.strip()

    # Deduplicated brand-only lookup: brand_lower -> reason
    brand_only_map = {}
    for entry in brand_only:
        b = entry["brand_lower"]
        if b not in brand_only_map:
            brand_only_map[b] = entry["reason"]

    # Sort full products longest name first so the most specific match wins
    full_products_sorted = sorted(full_products, key=lambda e: len(e["full_name_lower"]), reverse=True)
    meaningful_brands = sorted(
        {e["brand_lower"] for e in brand_only if len(e["brand_lower"]) >= 4},
        key=len, reverse=True
    )

    # Precompile the per-brand match patterns once (was previously recompiled
    # for every product row × every brand — O(rows * brands) re.compile calls).
    _brand_patterns = {}
    for b in meaningful_brands:
        pat = re.compile(r'^' + re.escape(b) + r'(\b|$)', re.IGNORECASE)
        name_pat = re.compile(r'^' + re.escape(b) + r'\b', re.IGNORECASE) if len(b) >= 4 else None
        _brand_patterns[b] = (pat, name_pat)

    # Many KEBS brand entries are short, common words fragmented out of a
    # longer real name — e.g. "soft" from "Soft & Beautiful", "fair" from
    # "Fair & White", "dream", "secret", "claire", "island". Matching those
    # against NAME unconditionally (the old behavior) flags completely
    # unrelated, legitimate products — "Nivea Soft Body Lotion" gets read as
    # the banned brand "Soft" even though BRAND clearly says Nivea. The
    # NAME-based fallback is only trustworthy when the seller's declared
    # BRAND is itself generic/evasive (the same gate check_suspected_fake_perfume
    # and the sneaker/jersey counterfeit checks already use before trusting a
    # brand word found only in NAME) — otherwise the declared BRAND field is
    # definitive and should win.
    _KEBS_GENERIC_BRAND_MARKERS = {
        "", "nan", "none", "generic", "fashion", "unbranded", "no brand",
        "original", "new", "designers collection", "smart collection",
    }
    target["_brand_is_generic"] = target["_brand_l"].isin(_KEBS_GENERIC_BRAND_MARKERS)

    # Vectorized pre-filter: two combined-regex scans (C speed) find the rows
    # that could possibly match, so the Python loop below — which preserves the
    # exact longest-name-first / reason-selection semantics — only runs on that
    # (usually tiny) candidate subset instead of every H&B row.
    _cand_mask = pd.Series(False, index=target.index)
    _full_name_pattern = "|".join(
        re.escape(e["full_name_lower"]) for e in full_products_sorted if len(e["full_name_lower"]) >= 5
    )
    if _full_name_pattern:
        # A substring of NAME is always a substring of _combined (brand + " " + name),
        # so scanning _combined covers both Step-1 checks in the loop.
        _cand_mask |= target["_combined"].str.contains(_full_name_pattern, regex=True, na=False)
    if meaningful_brands:
        _brand_start_pattern = r'^(?:' + "|".join(re.escape(b) for b in meaningful_brands) + r')(?:\b|$)'
        _cand_mask |= (
            (
                target["_brand_l"].str.contains(_brand_start_pattern, regex=True, na=False)
                | (
                    target["_brand_is_generic"]
                    & target["_name_l"].str.contains(_brand_start_pattern, regex=True, na=False)
                )
            )
            & target["_name_l"].str.contains(_SKIN_PRODUCT_TYPES, na=False)
        )
    target = target[_cand_mask]
    if target.empty:
        return pd.DataFrame(columns=data.columns)

    results = []
    seen_sids = set()

    for idx, row in target.iterrows():
        sid = str(row.get("PRODUCT_SET_SID", "")).strip()
        if sid in seen_sids:
            continue

        brand_val  = row["_brand_l"]
        name_val   = row["_name_l"]
        combined   = row["_combined"]

        matched_reason = None

        # ── Step 1: Full product name match ──────────────────────────────────
        # Check if any full KEBS product name is a substring of the combined text.
        for entry in full_products_sorted:
            kebs_name = entry["full_name_lower"]  # e.g. 'elegance skin lightening cream'
            if len(kebs_name) < 5:
                continue
            if kebs_name in combined:
                matched_reason = entry["reason"]
                break
            # Also check name alone (in case brand was in BRAND field)
            if kebs_name in name_val:
                matched_reason = entry["reason"]
                break

        # ── Step 2: Brand-only match + skin-product type check ───────────────
        # Only runs if Step 1 didn't already match.
        if not matched_reason:
            for b in meaningful_brands:
                # BRAND field must start with or equal the banned brand
                pat, name_pat = _brand_patterns[b]
                brand_match = pat.match(brand_val)

                # OR the name starts with the banned brand word — only trusted
                # when the seller's declared BRAND is itself generic/evasive
                # (see _KEBS_GENERIC_BRAND_MARKERS above). A specific declared
                # BRAND is definitive: "Nivea Soft Body Lotion" is Nivea, not
                # the banned brand "Soft", no matter what NAME says.
                name_brand_match = False
                if not brand_match and name_pat is not None and row["_brand_is_generic"]:
                    name_brand_match = bool(name_pat.match(name_val))

                if brand_match or name_brand_match:
                    # Confirm it looks like a skin product
                    if _SKIN_PRODUCT_TYPES.search(name_val):
                        matched_reason = brand_only_map.get(b, "Banned by KEBS.")
                        break

        if not matched_reason:
            continue

        seen_sids.add(sid)
        comment = (
            f"{matched_reason} Please contact Jumia Seller Support and raise a claim to confirm "
            f"whether this product is eligible for listing. See https://www.kebs.org/banned-products/"
        )
        result_row = row.copy()
        result_row["FLAG"] = "1000033 - Keywords in your content/ Product name / description has been blacklisted"
        result_row["Comment_Detail"] = comment
        results.append(result_row)

    if not results:
        return pd.DataFrame(columns=data.columns)

    return pd.DataFrame(results)




def apply_kebs_banned_rule(sid: str, rec: dict, has_status_col: bool, status: str, reason: str, base_row_fn) -> dict | None:
    is_rejection_like = (
        (has_status_col and status in ("rejected", "review", "manual review"))
        or (not has_status_col and reason and "error" not in reason.lower())
    )

    if is_rejection_like:
        return None

    full_products, brand_only = load_kebs_banned_products()
    if not full_products and not brand_only:
        return None

    hb_codes = load_kebs_hb_codes()
    if not hb_codes:
        return None

    cat_code = str(rec.get("CATEGORY_CODE", "")).strip()
    is_hb = (cat_code in hb_codes)
    
    if not is_hb:
        cat_str = str(rec.get("CATEGORY", "")).strip().lower()
        if cat_str:
            for allowed in _KEBS_ALLOWED_PATH_SECTIONS:
                if allowed.lower() in cat_str:
                    is_hb = True
                    break
                    
    if not is_hb:
        return None

    brand_val = str(rec.get("BRAND", "")).strip().lower()
    name_val  = str(rec.get("NAME", "")).strip().lower()
    combined  = (brand_val + " " + name_val).strip()

    brand_only_map = {}
    for entry in brand_only:
        b = entry["brand_lower"]
        if b not in brand_only_map:
            brand_only_map[b] = entry["reason"]

    full_products_sorted = sorted(full_products, key=lambda e: len(e["full_name_lower"]), reverse=True)
    meaningful_brands = sorted(
        {e["brand_lower"] for e in brand_only if len(e["brand_lower"]) >= 4},
        key=len, reverse=True
    )

    matched_reason = None

    # Step 1: full product name substring match
    for entry in full_products_sorted:
        kebs_name = entry["full_name_lower"]
        if len(kebs_name) < 5:
            continue
        if kebs_name in combined or kebs_name in name_val:
            matched_reason = entry["reason"]
            break

    _GENERIC_KEBS_BRANDS = {"skin", "body", "cream", "clear", "fade", "fair", "hot", "soft", "first", "dream"}

    # Precompile the per-brand match patterns once per call instead of per candidate brand.
    _brand_patterns = {}
    for b in meaningful_brands:
        if b in _GENERIC_KEBS_BRANDS:
            pat = re.compile(r'^' + re.escape(b) + r'$', re.IGNORECASE)
            name_pat = re.compile(r'^' + re.escape(b) + r'$', re.IGNORECASE)
        else:
            pat = re.compile(r'^' + re.escape(b) + r'(\b|$)', re.IGNORECASE)
            name_pat = re.compile(r'^' + re.escape(b) + r'\b', re.IGNORECASE)
        _brand_patterns[b] = (pat, name_pat)

    # Step 2: brand-only match + skin-type word
    if not matched_reason:
        for b in meaningful_brands:
            pat, name_pat = _brand_patterns[b]
            if pat.match(brand_val) or name_pat.match(name_val):
                if _SKIN_PRODUCT_TYPES.search(name_val):
                    matched_reason = brand_only_map.get(b, "Banned by KEBS.")
                    break

    if not matched_reason:
        return None

    detail = (
        f"{matched_reason} Please contact Jumia Seller Support and raise a claim to confirm "
        f"whether this product is eligible for listing. See https://www.kebs.org/banned-products/"
    )
    row = base_row_fn(sid, "kebs", rec)
    row.update({
        "Reason Type": "KEBS Banned Products",
        "Verdict": "False Approval",
        "Detail": detail
    })
    return row



@st.cache_data(ttl=3600)
def load_kebs_fda_keywords():
    if not os.path.exists("kebs_banned_products (1).xlsx"):
        return None
    try:
        df = pd.read_excel("kebs_banned_products (1).xlsx", sheet_name="Sheet1")
        if "FDA keywords" not in df.columns:
            return None
        kws = df["FDA keywords"].dropna().astype(str).str.strip().tolist()
        kws = [k for k in kws if k and k.lower() not in ("nan", "none")]
        if not kws:
            return None
        kws = sorted(kws, key=len, reverse=True)
        pattern = re.compile(
            r"(?<!\w)(?:" + "|".join(re.escape(k) for k in kws) + r")(?!\w)",
            re.IGNORECASE,
        )
        return pattern
    except Exception:
        return None


def check_kebs_fda(data: pd.DataFrame) -> pd.DataFrame:
    required = {"PRODUCT_SET_SID", "NAME", "BRAND", "CATEGORY_CODE"}
    if data.empty or not required.issubset(data.columns):
        return pd.DataFrame(columns=data.columns)

    pattern = load_kebs_fda_keywords()
    if not pattern:
        return pd.DataFrame(columns=data.columns)

    hb_codes = load_kebs_hb_codes()
    if not hb_codes:
        return pd.DataFrame(columns=data.columns)

    d = data.copy()
    
    # Filter for H&B categories
    cat_mask = d["CATEGORY_CODE"].astype(str).str.strip().isin(hb_codes)
    target = d[cat_mask]
    if target.empty:
        return pd.DataFrame(columns=data.columns)

    text_to_search = (target["BRAND"].fillna("") + " " + target["NAME"].fillna("")).str.lower()
    match_mask = text_to_search.str.contains(pattern, na=False)
    
    flagged = target[match_mask].copy()
    if flagged.empty:
        return pd.DataFrame(columns=data.columns)

    def _build_comment(row_text: str) -> str:
        matches = pattern.findall(row_text)
        kw = matches[0].upper() if matches else "FDA Controlled Substance"
        return f"Product name or brand contains '{kw}' which requires FDA approval. Please contact Jumia Seller Support to confirm eligibility."

    flagged["FLAG"] = "FDA"
    flagged["Comment_Detail"] = text_to_search[match_mask].apply(_build_comment)
    
    return flagged.drop_duplicates(subset=["PRODUCT_SET_SID"])


